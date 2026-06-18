"""
Evaluation — mc-classifier, senza ground truth.

Differenze rispetto a run_evaluationV2.py:
  - Agente target : mc-classifier (non orchestrator-agent-v2).
  - No ground truth: non esiste una risposta attesa fornita dal cliente.
    Il judge valuta coerenza, completezza e rilevanza basandosi solo su:
      (a) la query originale
      (b) la risposta testuale dell'agente
      (c) l'output strutturato estratto (classificazione + candidati)
  - Dataset embedded: le domande reali da "Domande Chatbot.doc" sono incluse
    direttamente nel file; `--dataset` accetta anche un .jsonl esterno.
  - Metriche adattate al ruolo del classifier:
      classification_coherence  — il JSON classificato rispecchia la query?
      field_completeness        — tutti i campi estraibili sono presenti?
      search_relevance          — i candidati restituiti sono pertinenti?
      candidate_quality         — i profili hanno requisiti adeguati?
      handling_complexity       — query vaghe/edge case gestite bene?
  - NO calibrazione dei punteggi: output grezzo del judge.
  - Output CSV + Excel (con openpyxl se disponibile).

Variabili d'ambiente richieste (o in local.settings.json):
  FOUNDRY_ENDPOINT      — es. https://foundry-ai-mc-dev.services.ai.azure.com
  FOUNDRY_PROJECT       — nome del progetto Foundry
  FOUNDRY_API_KEY       — chiave API per le Responses API
  FOUNDRY_MODEL         — es. gpt-4.1-mini  (modello dichiarato nell'agente)
  AZURE_OPENAI_ENDPOINT — endpoint Azure OpenAI per il judge
  AZURE_OPENAI_KEY      — chiave Azure OpenAI per il judge
  AZURE_OPENAI_MODEL    — deployment name per il judge (es. gpt-4.1-mini)

Usage:
    python scripts/run_evaluation_classifier.py
    python scripts/run_evaluation_classifier.py --max-rows 20
    python scripts/run_evaluation_classifier.py --output C:\\path\\results.csv
    python scripts/run_evaluation_classifier.py --dataset custom.jsonl
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import random
import sys
import time
import warnings
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Silence noisy deps
# ---------------------------------------------------------------------------
warnings.filterwarnings("ignore", message=r".[Pp]ydantic.", category=UserWarning)
warnings.filterwarnings("ignore", message=r".pydantic.", category=DeprecationWarning)
warnings.filterwarnings("ignore", message=r"Pydantic serializer warnings:.*", category=UserWarning)
try:
    from pydantic.warnings import PydanticDeprecatedSince20
    warnings.filterwarnings("ignore", category=PydanticDeprecatedSince20)
except Exception:
    pass

# ---------------------------------------------------------------------------
# Config — local.settings.json + env override
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO_ROOT   = _SCRIPT_DIR.parent

_settings_path = _REPO_ROOT / "local.settings.json"
_env: dict[str, str] = {}
if _settings_path.exists():
    try:
        _raw = json.loads(_settings_path.read_text(encoding="utf-8"))
        _env = _raw.get("Values", {})
    except Exception as exc:
        print(f"[WARNING] Could not parse local.settings.json: {exc}", file=sys.stderr)


def _cfg(key: str, default: str = "") -> str:
    """Legge prima da os.environ, poi da local.settings.json, poi usa default."""
    return os.environ.get(key) or _env.get(key) or default


# Foundry — Responses API con agent_reference
# PROJECT_ENDPOINT = https://<account>.services.ai.azure.com/api/projects/<project>
FOUNDRY_ENDPOINT  = _cfg("FOUNDRY_ENDPOINT", "https://foundry-ai-mc-dev.services.ai.azure.com/").rstrip("/")
FOUNDRY_PROJECT   = _cfg("FOUNDRY_PROJECT",  "test-project")
FOUNDRY_API_KEY   = _cfg("FOUNDRY_API_KEY",  "5Gum7Js3kot14QDeU2sbhi1THB83kVveBp9BkH635tV6JoGJIEPtJQQJ99CBACfhMk5XJ3w3AAAAACOGPvCI") or _cfg("AZURE_OPENAI_KEY", "")
FOUNDRY_API_VER   = _cfg("FOUNDRY_API_VERSION", "2025-05-15-preview")
AGENT_ID          = _cfg("CLASSIFIER_AGENT_ID", "mc-classifier")
FOUNDRY_MODEL     = _cfg("FOUNDRY_MODEL", _cfg("AZURE_OPENAI_MODEL", "gpt-4.1-mini"))

if FOUNDRY_PROJECT:
    PROJECT_ENDPOINT = f"{FOUNDRY_ENDPOINT}/api/projects/{FOUNDRY_PROJECT}"
else:
    # FOUNDRY_ENDPOINT può già essere il project endpoint completo
    PROJECT_ENDPOINT = FOUNDRY_ENDPOINT

# Judge — Azure OpenAI diretto
OAI_ENDPOINT = _cfg("AZURE_OPENAI_ENDPOINT", FOUNDRY_ENDPOINT).rstrip("/")
OAI_API_KEY  = _cfg("AZURE_OPENAI_KEY", FOUNDRY_API_KEY)
OAI_API_VER  = _cfg("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")
JUDGE_MODEL  = _cfg("JUDGE_DEPLOYMENT", _cfg("AZURE_OPENAI_MODEL", "gpt-4.1-mini"))

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
parser = argparse.ArgumentParser(
    description="Evaluation mc-classifier — senza ground truth."
)
parser.add_argument("--dataset",               default="")
parser.add_argument("--max-rows",              type=int,   default=0)
parser.add_argument("--output",                default="")
parser.add_argument("--agent-min-interval",    type=float, default=1.5)
parser.add_argument("--judge-min-interval",    type=float, default=1.0)
parser.add_argument("--max-retries",           type=int,   default=5)
parser.add_argument("--retry-base-seconds",    type=float, default=2.0)
parser.add_argument("--retry-max-seconds",     type=float, default=30.0)
parser.add_argument("--retry-jitter-seconds",  type=float, default=0.7)
args = parser.parse_args()

# ---------------------------------------------------------------------------
# Embedded dataset — domande da "Domande Chatbot.doc"
# Formato: {"title": "<categoria>", "query": "<domanda originale>"}
# Nessuna ground truth: il judge valuta senza risposta attesa.
# ---------------------------------------------------------------------------
_EMBEDDED_DATASET: list[dict[str, Any]] = [
    # ── Java Developer ──────────────────────────────────────────────────────
    {"title": "Java", "query": "Cerco uno sviluppatore Java senior con esperienza Spring Boot e microservizi per progetto bancario a Milano."},
    {"title": "Java", "query": "Avete profili Java disponibili da settembre in full remote?"},
    {"title": "Java", "query": "Mi serve un backend developer con Kafka e Kubernetes, almeno 4 anni di esperienza."},
    {"title": "Java", "query": "Cerchiamo un Java developer che abbia lavorato su architetture event driven."},
    {"title": "Java", "query": "Profilo Java + Angular full stack zona Torino ibrido."},
    {"title": "Java", "query": "Avete qualcuno con esperienza Quarkus?"},
    {"title": "Java", "query": "Cerco uno sviluppatore backend con Java 17 e Docker."},
    {"title": "Java", "query": "Team già formato su stack Java/Spring?"},
    {"title": "Java", "query": "Mi servirebbe una figura middle Java con inglese fluente per cliente estero."},
    {"title": "Java", "query": "Java developer con esperienza settore assicurativo."},
    {"title": "Java", "query": "Cerco risorsa urgente entro 2 settimane, stack Java + AWS."},
    {"title": "Java", "query": "Avete profili con esperienza in migrazione monolite-microservizi?"},
    {"title": "Java", "query": "Java developer disponibile part-time 3 giorni a settimana."},
    {"title": "Java", "query": "Mi serve un consulente Java onsite a Roma almeno 4 giorni."},
    {"title": "Java", "query": "Cerco profilo backend ma non troppo senior, budget contenuto."},
    # ── Frontend ────────────────────────────────────────────────────────────
    {"title": "Frontend", "query": "React developer senior con Typescript e Redux."},
    {"title": "Frontend", "query": "Cerco frontend Angular disponibile subito."},
    {"title": "Frontend", "query": "Avete qualcuno con esperienza Next.js?"},
    {"title": "Frontend", "query": "Frontend developer con competenze UX/UI."},
    {"title": "Frontend", "query": "Cerco sviluppatore React Native per progetto mobile."},
    {"title": "Frontend", "query": "Angular 16 + RxJS + ngrx."},
    {"title": "Frontend", "query": "Mi serve una figura frontend che sappia anche fare test Cypress."},
    {"title": "Frontend", "query": "Cerco sviluppatore Vue.js ma con disponibilità immediata."},
    {"title": "Frontend", "query": "Frontend developer con inglese C1."},
    {"title": "Frontend", "query": "Team frontend già strutturato su React?"},
    # ── Full Stack ───────────────────────────────────────────────────────────
    {"title": "FullStack", "query": "Full stack Java + Angular zona Bologna."},
    {"title": "FullStack", "query": "Cerco un full stack Node.js + React."},
    {"title": "FullStack", "query": "Avete sviluppatori .NET + Vue?"},
    {"title": "FullStack", "query": "Full stack con esperienza cloud AWS."},
    {"title": "FullStack", "query": "Cerco figura versatile che possa seguire sia frontend che backend."},
    {"title": "FullStack", "query": "Full stack senior per startup fintech."},
    {"title": "FullStack", "query": "Mi serve qualcuno autonomo lato sviluppo end-to-end."},
    {"title": "FullStack", "query": "Profilo full remote con stack MERN."},
    {"title": "FullStack", "query": "Cerco sviluppatore con competenze DevOps oltre al coding."},
    {"title": "FullStack", "query": "Full stack disponibile per progetto breve di 3 mesi."},
    # ── DevOps / Cloud ───────────────────────────────────────────────────────
    {"title": "DevOps", "query": "Cerco DevOps engineer con Kubernetes e Terraform."},
    {"title": "DevOps", "query": "Avete profili AWS certificati?"},
    {"title": "DevOps", "query": "Mi serve un cloud architect Azure."},
    {"title": "DevOps", "query": "DevOps con esperienza CI/CD GitLab."},
    {"title": "DevOps", "query": "Cerco sistemista cloud con esperienza Docker."},
    {"title": "DevOps", "query": "Figura SRE disponibile in tempi brevi."},
    {"title": "DevOps", "query": "Esperienza OpenShift obbligatoria."},
    {"title": "DevOps", "query": "Cerco consulente per setup infrastruttura AWS."},
    {"title": "DevOps", "query": "Kubernetes specialist onsite Milano."},
    {"title": "DevOps", "query": "DevOps junior con voglia di crescere."},
    # ── Data / AI ────────────────────────────────────────────────────────────
    {"title": "Data/AI", "query": "Cerco Data Engineer con Spark e Python."},
    {"title": "Data/AI", "query": "Avete data scientist con esperienza NLP?"},
    {"title": "Data/AI", "query": "Machine learning engineer con esperienza LLM."},
    {"title": "Data/AI", "query": "Cerco Power BI specialist."},
    {"title": "Data/AI", "query": "Profilo Tableau senior."},
    {"title": "Data/AI", "query": "Data analyst con SQL avanzato."},
    {"title": "Data/AI", "query": "Cerco figura AI Generativa per chatbot enterprise."},
    {"title": "Data/AI", "query": "Esperienza Databricks richiesta."},
    {"title": "Data/AI", "query": "Big Data engineer Hadoop/Kafka."},
    {"title": "Data/AI", "query": "Consulente BI settore finance."},
    # ── Cybersecurity ────────────────────────────────────────────────────────
    {"title": "Cybersecurity", "query": "Cerco cyber security analyst SOC."},
    {"title": "Cybersecurity", "query": "Ethical hacker disponibile onsite?"},
    {"title": "Cybersecurity", "query": "Esperienza IAM e CyberArk."},
    {"title": "Cybersecurity", "query": "Security engineer con ISO27001."},
    {"title": "Cybersecurity", "query": "Cerco penetration tester freelance."},
    {"title": "Cybersecurity", "query": "SOC analyst h24."},
    {"title": "Cybersecurity", "query": "Esperienza SIEM Splunk obbligatoria."},
    {"title": "Cybersecurity", "query": "Consulente GRC cybersecurity."},
    {"title": "Cybersecurity", "query": "Security architect cloud Azure."},
    {"title": "Cybersecurity", "query": "Cerco figura con clearance o esperienza ambito difesa."},
    # ── SAP / ERP ────────────────────────────────────────────────────────────
    {"title": "SAP/ERP", "query": "SAP FI senior per rollout internazionale."},
    {"title": "SAP/ERP", "query": "Cerco consulente SAP MM disponibile subito."},
    {"title": "SAP/ERP", "query": "SAP ABAP con esperienza S/4HANA."},
    {"title": "SAP/ERP", "query": "Avete profili SAP SD bilingue inglese/francese?"},
    {"title": "SAP/ERP", "query": "SAP consultant con esperienza manufacturing."},
    {"title": "SAP/ERP", "query": "Oracle ERP specialist."},
    {"title": "SAP/ERP", "query": "Dynamics 365 consultant."},
    {"title": "SAP/ERP", "query": "SAP PP onsite Torino."},
    {"title": "SAP/ERP", "query": "Cerco profilo SAP Basis."},
    {"title": "SAP/ERP", "query": "Team SAP già disponibile?"},
    # ── QA / Testing ─────────────────────────────────────────────────────────
    {"title": "QA/Testing", "query": "Test automation engineer Selenium."},
    {"title": "QA/Testing", "query": "QA manuale con esperienza finance."},
    {"title": "QA/Testing", "query": "Cerco automation tester Cypress."},
    {"title": "QA/Testing", "query": "Performance tester JMeter."},
    {"title": "QA/Testing", "query": "QA lead per coordinamento team."},
    {"title": "QA/Testing", "query": "Test engineer con API testing."},
    {"title": "QA/Testing", "query": "Esperienza test mobile richiesta."},
    {"title": "QA/Testing", "query": "QA con inglese fluente."},
    {"title": "QA/Testing", "query": "Tester disponibile solo part-time?"},
    {"title": "QA/Testing", "query": "Cerco figura junior QA."},
    # ── PM / Agile ───────────────────────────────────────────────────────────
    {"title": "PM/Agile", "query": "Cerco Scrum Master certificato."},
    {"title": "PM/Agile", "query": "PM IT con esperienza bancaria."},
    {"title": "PM/Agile", "query": "Agile coach disponibile da ottobre."},
    {"title": "PM/Agile", "query": "Project manager SAP rollout."},
    {"title": "PM/Agile", "query": "PM tecnico con background sviluppo."},
    {"title": "PM/Agile", "query": "Delivery manager per coordinamento fornitori."},
    {"title": "PM/Agile", "query": "Cerco PM con ottimo inglese."},
    {"title": "PM/Agile", "query": "Program manager ambito digital transformation."},
    {"title": "PM/Agile", "query": "PMO junior."},
    {"title": "PM/Agile", "query": "Cerco figura ibrida PM + analista funzionale."},
    # ── Business Analyst ─────────────────────────────────────────────────────
    {"title": "BusinessAnalyst", "query": "Business analyst settore assicurativo."},
    {"title": "BusinessAnalyst", "query": "Analista funzionale con esperienza CRM."},
    {"title": "BusinessAnalyst", "query": "Cerco functional analyst SAP."},
    {"title": "BusinessAnalyst", "query": "BA con capacità di raccolta requisiti."},
    {"title": "BusinessAnalyst", "query": "Esperienza UML e BPMN richiesta."},
    {"title": "BusinessAnalyst", "query": "Analista tecnico-funzionale payments."},
    {"title": "BusinessAnalyst", "query": "Functional analyst con SQL."},
    {"title": "BusinessAnalyst", "query": "Cerco figura ponte business-IT."},
    {"title": "BusinessAnalyst", "query": "Business analyst disponibile onsite Roma."},
    {"title": "BusinessAnalyst", "query": "Esperienza Agile obbligatoria."},
    # ── Vago / Generico ───────────────────────────────────────────────────────
    {"title": "Vago", "query": "Mi serve qualcuno bravo su cloud."},
    {"title": "Vago", "query": "Avete profili AI?"},
    {"title": "Vago", "query": "Cerco uno sviluppatore senior ma con costo contenuto."},
    {"title": "Vago", "query": "Mi serve una figura tecnica ma anche gestionale."},
    {"title": "Vago", "query": "Qualcuno disponibile subito per cliente enterprise."},
    {"title": "Vago", "query": "Cerco risorsa con esperienza moderna."},
    {"title": "Vago", "query": "Mi serve un profilo forte lato backend."},
    {"title": "Vago", "query": "Avete qualcuno che conosca bene Microsoft?"},
    {"title": "Vago", "query": "Cerco consulente esperto fintech."},
    {"title": "Vago", "query": "Ho bisogno di un profilo autonomo."},
    # ── Colloquiale / Abbreviato ──────────────────────────────────────────────
    {"title": "Colloquiale", "query": "Cerco urgentemente java senior milano ibrido banca."},
    {"title": "Colloquiale", "query": "fullstack react node remoto asap."},
    {"title": "Colloquiale", "query": "avete cv python ai disponibili?"},
    {"title": "Colloquiale", "query": "servirebbe devops kubernetes inglese ok."},
    {"title": "Colloquiale", "query": "candidato .net no junior."},
    {"title": "Colloquiale", "query": "ci serve una persona che possa interfacciarsi col cliente e fare sviluppo."},
    {"title": "Colloquiale", "query": "React o Angular va bene uguale."},
    {"title": "Colloquiale", "query": "Cerco figura tecnica ma che sappia parlare con il business."},
    {"title": "Colloquiale", "query": "Mi servirebbe qualcuno disponibile entro lunedì."},
    {"title": "Colloquiale", "query": "Team di 2-3 persone su Java riuscite a proporlo?"},
    # ── Edge Case ─────────────────────────────────────────────────────────────
    {"title": "EdgeCase", "query": "Cerco COBOL developer under 30."},
    {"title": "EdgeCase", "query": "Mi serve un DevOps con esperienza sia AWS che mainframe."},
    {"title": "EdgeCase", "query": "Full remote ma residente entro 50km da Padova."},
    {"title": "EdgeCase", "query": "Cerco AI engineer con esperienza sanitaria e conoscenza normativa GDPR."},
    {"title": "EdgeCase", "query": "SAP consultant disponibile solo mattina."},
    {"title": "EdgeCase", "query": "Profilo bilingue italiano/tedesco con esperienza banking."},
    {"title": "EdgeCase", "query": "Cerco sviluppatore React senior con budget da junior."},
    {"title": "EdgeCase", "query": "Mi serve una figura che sappia fare tutto."},
    {"title": "EdgeCase", "query": "Cerco consulente blockchain con esperienza assicurativa."},
    {"title": "EdgeCase", "query": "Avete qualcuno disponibile domani?"},
    # ── Consulenziale / Architetturale ────────────────────────────────────────
    {"title": "Consulenziale", "query": "Chi avete disponibile su Roma lato backend?"},
    {"title": "Consulenziale", "query": "Mi consigli un profilo DevOps senior?"},
    {"title": "Consulenziale", "query": "Abbiamo bisogno di rinforzare il team frontend."},
    {"title": "Consulenziale", "query": "Cerco una figura più architetturale che operativa."},
    {"title": "Consulenziale", "query": "Avete qualcuno forte su dati e AI?"},
    {"title": "Consulenziale", "query": "Mi serve una persona già abituata a contesti enterprise."},
    {"title": "Consulenziale", "query": "Qualcuno con esperienza startup ma strutturato."},
    {"title": "Consulenziale", "query": "Cerco sviluppatore ma con attitudine consulenziale."},
    {"title": "Consulenziale", "query": "Avete persone che parlano bene inglese?"},
    {"title": "Consulenziale", "query": "Chi potrei inserire su progetto cloud migration?"},
    # ── Multi-requisito / Complesso ───────────────────────────────────────────
    {"title": "MultiRequisito", "query": "Java senior + AWS + inglese + Milano ibrido + disponibilità entro 15 giorni."},
    {"title": "MultiRequisito", "query": "React developer full remote ma disponibile a trasferte mensili."},
    {"title": "MultiRequisito", "query": "Data engineer Python/Spark con esperienza finance e inglese fluente."},
    {"title": "MultiRequisito", "query": "PM SAP con esperienza internazionale e presenza onsite."},
    {"title": "MultiRequisito", "query": "Cybersecurity specialist con reperibilità weekend."},
    {"title": "MultiRequisito", "query": "Cerco una persona che possa seguire integrazioni API con sistemi esterni."},
    {"title": "MultiRequisito", "query": "Mi serve una figura molto autonoma lato tecnico."},
    {"title": "MultiRequisito", "query": "Cerco qualcuno che abbia già lavorato in contesti enterprise complessi."},
    {"title": "MultiRequisito", "query": "Cerco una persona forte lato performance applicative."},
    {"title": "MultiRequisito", "query": "Mi serve supporto per modernizzazione applicativa."},
    # ── Settoriale / Normativo ────────────────────────────────────────────────
    {"title": "Settoriale", "query": "Figura con esperienza assicurativa e normativa IVASS."},
    {"title": "Settoriale", "query": "Cerco qualcuno abituato a lavorare in Agile strutturato."},
    {"title": "Settoriale", "query": "Mi serve una persona molto pratica e poco teorica."},
    {"title": "Settoriale", "query": "Cerco qualcuno forte lato troubleshooting."},
    {"title": "Settoriale", "query": "Profilo tecnico ma con attitudine manageriale."},
    {"title": "Settoriale", "query": "Mi serve una figura che possa coordinare altri consulenti."},
    {"title": "Settoriale", "query": "Cerco supporto per attività di refactoring."},
    {"title": "Settoriale", "query": "Cerco figura con competenze DevSecOps."},
    {"title": "Settoriale", "query": "Mi serve supporto su CI/CD pipeline."},
    {"title": "Settoriale", "query": "Figura con esperienza GitLab e automazione deployment."},
    # ── Database / Infrastruttura ─────────────────────────────────────────────
    {"title": "Database", "query": "Cerco persona forte su database relazionali."},
    {"title": "Database", "query": "Esperienza PostgreSQL gradita."},
    {"title": "Database", "query": "Profilo con competenze NoSQL."},
    {"title": "Database", "query": "Mi serve qualcuno che abbia lavorato su MongoDB."},
    {"title": "Infrastructure", "query": "Cerco supporto per applicativi mission critical."},
    {"title": "Infrastructure", "query": "Figura con esperienza alta affidabilità sistemi."},
    {"title": "Infrastructure", "query": "Esperienza Linux obbligatoria."},
    {"title": "Infrastructure", "query": "Mi serve competenza Windows Server."},
    {"title": "Infrastructure", "query": "Cerco figura con esperienza VMware."},
    {"title": "Infrastructure", "query": "Profilo con competenze networking."},
]


# ---------------------------------------------------------------------------
# Dataset loader
# ---------------------------------------------------------------------------

def _text(value: Any) -> str:
    return str(value or "").strip()


def _fix_mojibake(s: str) -> str:
    markers = ("Ã", "Â", "â")
    if not any(m in s for m in markers):
        return s
    original_noise = sum(s.count(m) for m in markers)
    for enc in ("cp1252", "latin-1"):
        try:
            fixed = s.encode(enc).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        if sum(fixed.count(m) for m in markers) < original_noise:
            return fixed
    return s


def load_jsonl(path: Path, max_rows: int = 0) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8-sig") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                print(f"[WARNING] Skipping malformed line {lineno}: {exc}", file=sys.stderr)
                continue
            if not isinstance(row, dict):
                continue
            rows.append({k: _fix_mojibake(v) if isinstance(v, str) else v for k, v in row.items()})
            if max_rows > 0 and len(rows) >= max_rows:
                break
    return rows


# ---------------------------------------------------------------------------
# Resolve dataset
# ---------------------------------------------------------------------------
if args.dataset:
    dataset_path = Path(args.dataset)
    if not dataset_path.exists():
        print(f"[ERROR] Dataset not found: {dataset_path}", file=sys.stderr)
        sys.exit(1)
    rows = load_jsonl(dataset_path, args.max_rows)
    print(f"[INFO] Loaded {len(rows)} rows from {dataset_path.name}")
else:
    rows = _EMBEDDED_DATASET
    if args.max_rows > 0:
        rows = rows[: args.max_rows]
    print(f"[INFO] Using embedded dataset ({len(rows)} domande)")

# ---------------------------------------------------------------------------
# Import dependencies
# ---------------------------------------------------------------------------
try:
    from openai import AzureOpenAI
except ImportError:
    print("[ERROR] Run: pip install openai", file=sys.stderr)
    sys.exit(1)

if not OAI_API_KEY:
    print("[ERROR] AZURE_OPENAI_KEY not set.", file=sys.stderr)
    sys.exit(1)
if not FOUNDRY_API_KEY:
    print("[ERROR] FOUNDRY_API_KEY (or AZURE_OPENAI_KEY) not set.", file=sys.stderr)
    sys.exit(1)
if not PROJECT_ENDPOINT:
    print(
        "[ERROR] FOUNDRY_ENDPOINT not set. "
        "Es: https://foundry-ai-mc-dev.services.ai.azure.com  +  FOUNDRY_PROJECT=<project-name>",
        file=sys.stderr,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Clients
# ---------------------------------------------------------------------------
oai_client = AzureOpenAI(
    azure_endpoint=OAI_ENDPOINT,
    api_version=OAI_API_VER,
    api_key=OAI_API_KEY,
)

agent_client = AzureOpenAI(
    azure_endpoint=PROJECT_ENDPOINT,
    api_version=FOUNDRY_API_VER,
    api_key=FOUNDRY_API_KEY,
)

# ---------------------------------------------------------------------------
# Metrics — adattate al ruolo del classifier
# ---------------------------------------------------------------------------
CORE_METRICS = [
    "classification_coherence",  # la classificazione strutturata rispecchia l'intento della query?
    "field_completeness",        # tutti i campi estraibili sono stati valorizzati?
    "search_relevance",          # i candidati restituiti sono pertinenti alla richiesta?
    "candidate_quality",         # i profili hanno i requisiti richiesti?
    "handling_complexity",       # query vaghe/ambigue/edge case gestite correttamente?
]

FINAL_SCORE_WEIGHTS = {
    "classification_coherence": 0.35,
    "search_relevance":         0.30,
    "field_completeness":       0.15,
    "candidate_quality":        0.15,
    "handling_complexity":      0.05,
}

# ---------------------------------------------------------------------------
# Judge prompts
# ---------------------------------------------------------------------------

_JUDGE_SYSTEM = """\
Sei un valutatore esperto di sistemi AI per il recruiting IT (staffing B2B).

Stai valutando le risposte del classificatore mc-classifier del sistema MC Flash.
Questo agente:
1. Interpreta una query in linguaggio naturale di un cliente che cerca un profilo IT
2. Estrae campi strutturati (skills, ruolo, location, seniority, lingua, work_mode, ecc.)
3. Chiama internamente un motore di ricerca per trovare candidati nel DB MC
4. Restituisce in testo naturale i candidati trovati con motivazioni

Non esiste una ground truth fornita dal cliente. Devi valutare:
  (a) Se la classificazione strutturata estratta è coerente con la query originale
  (b) Se i campi estraibili sono stati correttamente popolati
  (c) Se i candidati presentati sono pertinenti e ben selezionati
  (d) Se query difficili (vaghe, colloquiali, edge case) sono gestite bene

Principi fondamentali:
- Non penalizzare l'agente se il DB non contiene candidati perfetti (assenza ≠ errore agente).
- Penalizza se i candidati restituiti sono palesemente incompatibili con la query.
- Devi valutare anche l'evidenza del motore di ricerca (search index) quando disponibile,
    non solo il testo finale della risposta.
- Premia la corretta gestione della ambiguità: chiedere chiarimenti UNA domanda mirata
  quando la query ha meno di 2 segnali utili è comportamento corretto.
- Per query vaghe (es. "Avete profili AI?"), un tentativo di ricerca con i segnali
  disponibili è preferibile al blocco immediato.
- Sii determinista e restituisci solo JSON valido senza markdown.
"""

_JUDGE_PROMPT = """\
Valuta la risposta dell'agente mc-classifier per una richiesta
di ricerca profilo IT in italiano.

NON hai una ground truth. Valuta basandoti esclusivamente su:
1. La coerenza logica tra QUERY e risposta/classificazione
2. La completezza e correttezza dei campi estratti (se visibili nella risposta)
3. La pertinenza dei candidati presentati rispetto alla query
4. La qualità della gestione di query complesse, vaghe o edge case

Se è presente SEARCH INDEX EVIDENCE, usala come fonte primaria per valutare
search_relevance e candidate_quality. Il testo finale dell'agente è secondario.

Classifica la risposta in esattamente uno di:
- COHERENT_MATCH       : classificazione e candidati coerenti con la query
- COHERENT_NO_RESULTS  : classificazione corretta ma nessun candidato trovato (DB vuoto su quel profilo)
- PARTIAL_MATCH        : classificazione parzialmente corretta o candidati solo parzialmente pertinenti
- CLARIFICATION_ASKED  : agente ha chiesto chiarimenti (valuta se appropriato dato la query)
- INCOHERENT           : classificazione o candidati non coerenti con la query
- AGENT_ERROR          : agente ha restituito errore o risposta vuota/incomprensibile

Definizioni metriche:

classification_coherence (1-5):
  La classificazione strutturata interna (skills, ruolo, location, seniority, ecc.)
  rispecchia l'intento della query? Se non è visibile nella risposta, inferisci
  dalla risposta testuale se l'agente ha capito correttamente la richiesta.
  5 = perfettamente coerente; 1 = completamente incoerente o assente.

field_completeness (1-5):
  L'agente ha estratto e utilizzato tutti i campi estraibili dalla query?
  Es. se la query menziona location, seniority, skill specifiche → sono state catturate?
  Per query vaghe, è accettabile avere pochi campi valorizzati.
  5 = tutti i campi estraibili catturati; 1 = campi ovvi mancanti.

search_relevance (1-5):
  I candidati presentati (se presenti) corrispondono ai requisiti della query?
  Se nessun candidato è stato trovato: 3 (neutro, non imputabile all'agente).
  Se candidati presentati ma palesemente incompatibili: 1-2.
  5 = tutti i candidati altamente pertinenti.

candidate_quality (1-5):
  I profili presentati hanno i requisiti tecnici richiesti dalla query?
  Considera skills, ruolo, location, seniority menzionati nella risposta.
  Se nessun candidato: 3 (neutro).
  5 = profili con requisiti pienamente soddisfatti.

handling_complexity (1-5):
  Per query chiare e specifiche: 5 se risposta corretta, 3 se parziale, 1 se incoerente.
  Per query vaghe/colloquiali/edge case: premia tentativi di ricerca intelligenti
  e domande di chiarimento pertinenti. Penalizza blocchi immotivati o risposte standard.
  5 = gestione eccellente del tipo di query; 1 = gestione completamente inadeguata.

Scala punteggi:
  5 = eccellente (raro)
  4 = buono con piccoli margini
  3 = accettabile / neutro
  2 = problematico
  1 = fallimento grave

Restituisci esattamente questo JSON:
{
  "classification": "COHERENT_MATCH",
  "classification_coherence": 1,
  "field_completeness": 1,
  "search_relevance": 1,
  "candidate_quality": 1,
  "handling_complexity": 1,
  "final_score": 3.0,
  "reasoning": {
    "classification_coherence": "...",
    "field_completeness": "...",
    "search_relevance": "...",
    "candidate_quality": "...",
    "handling_complexity": "..."
  },
  "classifier_strengths": "...",
  "classifier_weaknesses": "...",
  "improvement_suggestion": "..."
}

CATEGORIA QUERY: _TITLE_

QUERY ORIGINALE:
_QUERY_

CLASSIFICAZIONE STRUTTURATA ESTRATTA:
_STRUCTURED_CLASSIFICATION_

SEARCH INDEX EVIDENCE:
_SEARCH_EVIDENCE_

RISPOSTA AGENTE:
_RESPONSE_
"""

# ---------------------------------------------------------------------------
# Query evaluator prompts — diagnostica indipendente dalla risposta
# ---------------------------------------------------------------------------

_QUERY_EVAL_SYSTEM = """\
Sei un esperto valutatore di query per sistemi di matching IT B2B (staffing).
Il tuo compito è diagnosticare la qualità della domanda di un cliente
per capire se un eventuale punteggio basso dipende dalla query o dall'agente.
Rispondi solo con JSON valido, senza markdown.
"""

_QUERY_EVAL_PROMPT = """\
Valuta la qualità di questa QUERY di ricerca profilo IT.

Definizioni:
- query_clarity (1-5): La query è grammaticalmente chiara e non ambigua?
  5=chiarissima; 1=incomprensibile o fortemente ambigua.
- query_complexity (SIMPLE/MODERATE/COMPLEX): Quanta elaborazione richiede?
  SIMPLE=lookup diretto; MODERATE=ragionamento singolo; COMPLEX=multi-criterio o multi-step.
- query_specificity (HIGH/MEDIUM/LOW): Quanto è specifica la richiesta?
  HIGH=skills e ruolo precisi; LOW=richiesta generica senza dettagli tecnici.
- query_extractability (1-5): Quanto è facile estrarre campi strutturati (skills, ruolo, location)?
  5=tutti i campi espliciti; 1=nessun segnale strutturabile.
- query_root_cause: Dato quanto sopra, causa più probabile di un errore:
  QUESTION  = la query è ambigua, incompleta o fuorviante
  AGENT     = la query è chiara, un agente capace dovrebbe rispondere correttamente
  NONE      = query eccellente, errori improbabili
- query_eval_notes: Una o due frasi di diagnosi (italiano o inglese).

Restituisci esattamente:
{
  "query_clarity": 1,
  "query_complexity": "SIMPLE",
  "query_specificity": "HIGH",
  "query_extractability": 1,
  "query_root_cause": "AGENT",
  "query_eval_notes": "..."
}

CATEGORIA: _TITLE_
QUERY: _QUERY_
"""

QUERY_EVAL_METRICS = [
    "query_clarity",
    "query_complexity",
    "query_specificity",
    "query_extractability",
    "query_root_cause",
    "query_eval_notes",
]

# ---------------------------------------------------------------------------
# Retry / throttle helpers
# ---------------------------------------------------------------------------
_LAST_CALL_TS: dict[str, float] = {"agent": 0.0, "judge": 0.0}


def _wait_channel(channel: str, min_interval: float) -> None:
    if min_interval <= 0:
        return
    elapsed = time.time() - _LAST_CALL_TS.get(channel, 0.0)
    wait = min_interval - elapsed
    if wait > 0:
        time.sleep(wait)
    _LAST_CALL_TS[channel] = time.time()


def _is_rate_limit(exc: Exception) -> bool:
    status = getattr(exc, "status_code", None) or getattr(getattr(exc, "response", None), "status_code", None)
    if status == 429:
        return True
    return "429" in str(exc).lower() or "rate_limit" in str(exc).lower()


def _backoff(exc: Exception, attempt: int) -> float:
    header = None
    resp = getattr(exc, "response", None)
    if resp is not None:
        hdrs = getattr(resp, "headers", None) or {}
        val = hdrs.get("retry-after") or hdrs.get("Retry-After")
        if val:
            try:
                header = float(val)
            except (TypeError, ValueError):
                pass
    base = header if header is not None else min(
        args.retry_base_seconds * (2 ** max(attempt - 1, 0)),
        args.retry_max_seconds,
    )
    return base + random.uniform(0.0, args.retry_jitter_seconds)


def _call_retry(fn: Any, *, label: str, channel: str, min_interval: float) -> Any:
    retries = max(0, args.max_retries)
    attempt = 0
    while True:
        _wait_channel(channel, min_interval)
        try:
            return fn()
        except Exception as exc:
            if not _is_rate_limit(exc):
                raise
            if attempt >= retries:
                raise
            attempt += 1
            sleep_s = _backoff(exc, attempt)
            print(f" [THROTTLE] {label} retry {attempt}/{retries} in {sleep_s:.1f}s", end="", flush=True)
            time.sleep(sleep_s)


# ---------------------------------------------------------------------------
# Agent call — mc-classifier
# ---------------------------------------------------------------------------

def _response_to_dict(response: Any) -> dict[str, Any]:
    if response is None:
        return {}
    if hasattr(response, "model_dump"):
        try:
            dumped = response.model_dump()
            if isinstance(dumped, dict):
                return dumped
        except Exception:
            pass
    if isinstance(response, dict):
        return response
    try:
        return dict(vars(response))
    except Exception:
        return {}


def _extract_json_blocks(text: str) -> list[dict[str, Any]]:
    """Estrae tutti i blocchi JSON validi da una stringa."""
    results: list[dict[str, Any]] = []
    depth = 0
    start = -1
    for i, ch in enumerate(text):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start != -1:
                candidate = text[start : i + 1]
                try:
                    parsed = json.loads(candidate)
                    if isinstance(parsed, dict):
                        results.append(parsed)
                except json.JSONDecodeError:
                    pass
                start = -1
    return results


def _extract_search_index_evidence(response_dict: dict[str, Any]) -> dict[str, Any]:
    """Estrae evidenza search-index in modo robusto dal payload Responses API."""
    candidates: list[dict[str, Any]] = []

    def _looks_like_candidate(item: dict[str, Any]) -> bool:
        return any(k in item for k in ("name", "full_name", "candidate_id", "skills", "role", "location"))

    def _normalize_candidate(item: dict[str, Any]) -> dict[str, Any]:
        return {
            "name": _text(item.get("name") or item.get("full_name") or item.get("candidate_id")),
            "role": _text(item.get("role")),
            "location": _text(item.get("location")),
            "skills": item.get("skills") if isinstance(item.get("skills"), list) else [],
            "retrieval_score": item.get("retrieval_score"),
            "source_path": _text(item.get("source_path") or item.get("path") or item.get("source")),
        }

    def _to_mapping(value: Any) -> dict[str, Any] | None:
        if isinstance(value, dict):
            return value
        if hasattr(value, "model_dump"):
            try:
                dumped = value.model_dump()
                if isinstance(dumped, dict):
                    return dumped
            except Exception:
                pass
        try:
            dumped = dict(vars(value))
            if isinstance(dumped, dict):
                return dumped
        except Exception:
            pass
        return None

    def _walk(value: Any) -> None:
        mapping = _to_mapping(value)
        if mapping is not None and not isinstance(value, dict):
            value = mapping

        if isinstance(value, dict):
            for key in ("hits", "results"):
                maybe_list = value.get(key)
                if isinstance(maybe_list, list):
                    for obj in maybe_list:
                        if isinstance(obj, dict) and _looks_like_candidate(obj):
                            candidates.append(_normalize_candidate(obj))
            for child in value.values():
                _walk(child)
        elif isinstance(value, list):
            for child in value:
                _walk(child)
        elif isinstance(value, str):
            s = value.strip()
            if s.startswith("{") or s.startswith("["):
                try:
                    parsed = json.loads(s)
                    _walk(parsed)
                except json.JSONDecodeError:
                    pass

    _walk(response_dict)

    # Dedup by name+role+location
    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for c in candidates:
        key = "|".join([
            _text(c.get("name")).casefold(),
            _text(c.get("role")).casefold(),
            _text(c.get("location")).casefold(),
        ])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(c)

    top = deduped[:20]
    names = [str(c.get("name") or "") for c in top if str(c.get("name") or "")]
    evidence_payload = {
        "candidates_count": len(deduped),
        "top_candidates": top,
    }

    return {
        "candidates_count": len(deduped),
        "candidates_names": ", ".join(names[:5]),
        "search_index_evidence": json.dumps(evidence_payload, ensure_ascii=False),
    }


def _parse_classifier_output(response_text: str, response_dict: dict[str, Any]) -> dict[str, Any]:
    """
    Prova a estrarre i campi strutturati dalla risposta del classifier.
    Il classifier restituisce testo naturale; la classificazione strutturata
    potrebbe essere visibile nei tool call output interni.
    """
    result: dict[str, Any] = {
        "classified_skills":    "",
        "classified_role":      "",
        "classified_location":  "",
        "classified_seniority": "",
        "classified_language":  "",
        "classified_work_mode": "",
        "candidates_count":     0,
        "candidates_names":     "",
        "search_index_evidence": "",
    }

    def _merge_structured_fields(source: dict[str, Any]) -> None:
        skills = source.get("skills") or []
        if skills and isinstance(skills, list) and not result["classified_skills"]:
            result["classified_skills"] = ", ".join(str(s) for s in skills[:5])
        if not result["classified_role"]:
            result["classified_role"] = _text(source.get("role"))
        if not result["classified_location"]:
            result["classified_location"] = _text(source.get("location"))
        if not result["classified_seniority"]:
            result["classified_seniority"] = _text(source.get("seniority"))
        if not result["classified_language"]:
            result["classified_language"] = _text(source.get("language"))
        if not result["classified_work_mode"]:
            result["classified_work_mode"] = _text(source.get("work_mode"))

    def _walk_for_structured_fields(value: Any) -> None:
        if isinstance(value, dict):
            for key in ("search_request", "interpreted_request", "classification"):
                candidate = value.get(key)
                if isinstance(candidate, dict):
                    _merge_structured_fields(candidate)
            for child in value.values():
                _walk_for_structured_fields(child)
        elif isinstance(value, list):
            for child in value:
                _walk_for_structured_fields(child)

    # Cerca nei tool call output
    output_items = response_dict.get("output") or []
    if isinstance(output_items, list):
        for item in output_items:
            if not isinstance(item, dict):
                if hasattr(item, "model_dump"):
                    try:
                        item = item.model_dump()
                    except Exception:
                        continue
                else:
                    try:
                        item = dict(vars(item))
                    except Exception:
                        continue
            item_type = str(item.get("type") or "").lower()

            # Estrae i campi strutturati direttamente dagli arguments del tool call
            # (fonte piu' stabile rispetto al parsing della risposta testuale).
            if "call" in item_type:
                raw_args = item.get("arguments")
                if isinstance(raw_args, str) and raw_args.strip():
                    try:
                        parsed_args = json.loads(raw_args)
                    except json.JSONDecodeError:
                        parsed_args = None
                    if isinstance(parsed_args, dict):
                        _walk_for_structured_fields(parsed_args)

            # Tool call output contiene la risposta del searcher-wrapper
            if "call_output" in item_type or "output" in item_type:
                raw_output = item.get("output") or item.get("content") or ""
                candidate_payload = ""
                if isinstance(raw_output, str):
                    candidate_payload = raw_output
                elif isinstance(raw_output, dict):
                    # Foundry OpenAPI output tipico: {"response": "{...json...}"}
                    maybe_response = raw_output.get("response")
                    if isinstance(maybe_response, str):
                        candidate_payload = maybe_response
                    else:
                        candidate_payload = json.dumps(raw_output, ensure_ascii=False)

                if candidate_payload:
                    parsed_payload: Any | None = None
                    try:
                        parsed_payload = json.loads(candidate_payload)
                    except json.JSONDecodeError:
                        parsed_payload = None

                    if parsed_payload is not None:
                        _walk_for_structured_fields(parsed_payload)

                        # Conta candidati da search_response / hits / results
                        def _walk_for_hits(value: Any) -> None:
                            if isinstance(value, dict):
                                for key in ("hits", "results"):
                                    maybe = value.get(key)
                                    if isinstance(maybe, list) and maybe:
                                        result["candidates_count"] = max(result["candidates_count"], len(maybe))
                                        names = [
                                            str(h.get("name") or h.get("full_name") or h.get("candidate_id") or "")
                                            for h in maybe[:5]
                                            if isinstance(h, dict)
                                        ]
                                        joined = ", ".join(n for n in names if n)
                                        if joined:
                                            result["candidates_names"] = joined
                                for child in value.values():
                                    _walk_for_hits(child)
                            elif isinstance(value, list):
                                for child in value:
                                    _walk_for_hits(child)

                        _walk_for_hits(parsed_payload)
                    else:
                        # Fallback legacy parser se payload non e' JSON pulito
                        blocks = _extract_json_blocks(candidate_payload)
                        for block in blocks:
                            _walk_for_structured_fields(block)

    # Fallback: cerca JSON nella risposta testuale
    if not result["classified_skills"]:
        blocks = _extract_json_blocks(response_text)
        for block in blocks:
            _merge_structured_fields(block)

    # Fallback robusto: estrae evidenza search-index da tutto il payload response
    if result["candidates_count"] <= 0:
        evidence = _extract_search_index_evidence(response_dict)
        if evidence["candidates_count"] > 0:
            result["candidates_count"] = evidence["candidates_count"]
            result["candidates_names"] = evidence["candidates_names"]
            result["search_index_evidence"] = evidence["search_index_evidence"]
    else:
        result["search_index_evidence"] = json.dumps(
            {
                "candidates_count": result["candidates_count"],
                "top_candidates": result["candidates_names"],
            },
            ensure_ascii=False,
        )

    return result


def call_agent(query: str) -> dict[str, Any]:
    """Chiama mc-classifier via Foundry Responses API."""
    try:
        response = _call_retry(
            lambda: agent_client.responses.create(
                model=FOUNDRY_MODEL,
                input=query,
                extra_body={
                    "agent_reference": {
                        "name": AGENT_ID,
                        "type": "agent_reference",
                    }
                },
            ),
            label="agent",
            channel="agent",
            min_interval=args.agent_min_interval,
        )
        response_text = getattr(response, "output_text", "") or ""
        response_dict = _response_to_dict(response)
        response_id   = str(response_dict.get("id") or "")
        parsed        = _parse_classifier_output(response_text, response_dict)
        return {
            "response_text": response_text,
            "response_id":   response_id,
            "agent_error":   "",
            **parsed,
        }
    except Exception as exc:
        print(f"\n[AGENT ERROR] {exc}", file=sys.stderr)
        return {
            "response_text":        "",
            "response_id":          "",
            "agent_error":          str(exc),
            "classified_skills":    "",
            "classified_role":      "",
            "classified_location":  "",
            "classified_seniority": "",
            "classified_language":  "",
            "classified_work_mode": "",
            "candidates_count":     0,
            "candidates_names":     "",
            "search_index_evidence": "",
        }


# ---------------------------------------------------------------------------
# Judge — raw output, NO calibration
# ---------------------------------------------------------------------------

def _coerce_score(value: Any) -> int | None:
    if value is None:
        return None
    try:
        score = int(round(float(value)))
        return score if 1 <= score <= 5 else None
    except (TypeError, ValueError):
        return None


def _extract_json(raw: str) -> dict[str, Any] | None:
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        pass
    s, e = raw.find("{"), raw.rfind("}")
    if s == -1 or e <= s:
        return None
    try:
        parsed = json.loads(raw[s : e + 1])
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        return None


def _calculate_final_score(result: dict[str, Any]) -> float | None:
    total = 0.0
    for metric, weight in FINAL_SCORE_WEIGHTS.items():
        value = result.get(metric)
        if not isinstance(value, (int, float)):
            return None
        total += float(value) * weight
    return round(total, 2)


def judge_score(
    query: str,
    title: str,
    response: str,
    structured_classification: str,
    search_evidence: str,
) -> dict[str, Any]:
    """Chiama il judge e restituisce score RAW — nessuna calibrazione."""
    prompt = (
        _JUDGE_PROMPT
        .replace("_TITLE_", title)
        .replace("_QUERY_", query)
        .replace("_STRUCTURED_CLASSIFICATION_", structured_classification)
        .replace("_SEARCH_EVIDENCE_", search_evidence)
        .replace("_RESPONSE_", response)
    )
    empty: dict[str, Any] = {
        "classification": "PARTIAL_MATCH",
        **{m: None for m in CORE_METRICS},
        "final_score": None,
        "classifier_strengths":   "",
        "classifier_weaknesses":  "",
        "improvement_suggestion": "",
        **{f"reasoning_{m}": "" for m in CORE_METRICS},
    }
    try:
        resp = _call_retry(
            lambda: oai_client.chat.completions.create(
                model=JUDGE_MODEL,
                messages=[
                    {"role": "system", "content": _JUDGE_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                max_tokens=1000,
                temperature=0,
                response_format={"type": "json_object"},
            ),
            label="judge",
            channel="judge",
            min_interval=args.judge_min_interval,
        )
        raw = (resp.choices[0].message.content or "").strip()
    except Exception as exc:
        empty["reasoning_classification_coherence"] = f"Judge call failed: {exc}"
        return empty

    parsed = _extract_json(raw)
    if parsed is None:
        empty["reasoning_classification_coherence"] = "Judge returned malformed JSON."
        return empty

    result: dict[str, Any] = {
        "classification": _text(parsed.get("classification")).upper() or "PARTIAL_MATCH",
        **{m: _coerce_score(parsed.get(m)) for m in CORE_METRICS},
    }
    reasoning = parsed.get("reasoning", {}) or {}
    for m in CORE_METRICS:
        result[f"reasoning_{m}"] = _text(reasoning.get(m))

    result["classifier_strengths"]   = _text(parsed.get("classifier_strengths"))
    result["classifier_weaknesses"]  = _text(parsed.get("classifier_weaknesses"))
    result["improvement_suggestion"] = _text(parsed.get("improvement_suggestion"))

    judge_fs = parsed.get("final_score")
    try:
        fs = round(float(judge_fs), 2) if judge_fs is not None else None
    except (TypeError, ValueError):
        fs = None
    result["final_score"] = fs if fs is not None else _calculate_final_score(result)

    return result


# ---------------------------------------------------------------------------
# Query evaluator
# ---------------------------------------------------------------------------

def assess_query(query: str, title: str) -> dict[str, Any]:
    """Valuta la qualità della query in ingresso indipendentemente dalla risposta."""
    prompt = (
        _QUERY_EVAL_PROMPT
        .replace("_TITLE_", title)
        .replace("_QUERY_", query)
    )
    empty: dict[str, Any] = {
        "query_clarity":        None,
        "query_complexity":     "",
        "query_specificity":    "",
        "query_extractability": None,
        "query_root_cause":     "",
        "query_eval_notes":     "",
    }
    try:
        resp = _call_retry(
            lambda: oai_client.chat.completions.create(
                model=JUDGE_MODEL,
                messages=[
                    {"role": "system", "content": _QUERY_EVAL_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                max_tokens=400,
                temperature=0,
                response_format={"type": "json_object"},
            ),
            label="query_eval",
            channel="judge",
            min_interval=args.judge_min_interval,
        )
        raw = (resp.choices[0].message.content or "").strip()
    except Exception as exc:
        empty["query_eval_notes"] = f"Query eval failed: {exc}"
        return empty

    parsed = _extract_json(raw)
    if parsed is None:
        empty["query_eval_notes"] = "Query eval returned malformed JSON."
        return empty

    return {
        "query_clarity":        _coerce_score(parsed.get("query_clarity")),
        "query_complexity":     _text(parsed.get("query_complexity")).upper(),
        "query_specificity":    _text(parsed.get("query_specificity")).upper(),
        "query_extractability": _coerce_score(parsed.get("query_extractability")),
        "query_root_cause":     _text(parsed.get("query_root_cause")).upper(),
        "query_eval_notes":     _text(parsed.get("query_eval_notes")),
    }


# ---------------------------------------------------------------------------
# Output path
# ---------------------------------------------------------------------------
_ts = time.strftime("%Y%m%d%H%M%S")
output_path = Path(args.output) if args.output else Path(f"eval_classifier_{_ts}.csv")

# ---------------------------------------------------------------------------
# CSV columns
# ---------------------------------------------------------------------------
CSV_COLUMNS = [
    # Input
    "query", "title",
    # Risposta agente
    "response", "agent_error",
    # Campi estratti dal classifier
    "classified_skills", "classified_role", "classified_location",
    "classified_seniority", "classified_language", "classified_work_mode",
    "candidates_count", "candidates_names",
    "search_index_evidence",
    # Punteggi judge
    "classification", "final_score",
    *CORE_METRICS,
    *[f"reasoning_{m}" for m in CORE_METRICS],
    "classifier_strengths", "classifier_weaknesses", "improvement_suggestion",
    # Diagnostica query
    *QUERY_EVAL_METRICS,
    # Timing
    "latency_seconds", "judge_latency_seconds", "response_id",
]

# ---------------------------------------------------------------------------
# Excel helper
# ---------------------------------------------------------------------------

_SCORE_COLORS = {1: "FFCCCC", 2: "FFE5CC", 3: "FFFFCC", 4: "CCFFCC", 5: "99FF99"}
_CLASSIFICATION_COLORS = {
    "COHERENT_MATCH":      "55EE55",
    "COHERENT_NO_RESULTS": "AAFFAA",
    "PARTIAL_MATCH":       "FFEE44",
    "CLARIFICATION_ASKED": "B3D9FF",
    "INCOHERENT":          "FF9944",
    "AGENT_ERROR":         "FF4444",
}
_ROOT_CAUSE_COLORS = {
    "QUESTION": "FFB3B3",
    "AGENT":    "B3D9FF",
    "NONE":     "D5FFD5",
}


def _write_excel(results: list[dict[str, Any]], path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "[WARNING] openpyxl non installato — skip Excel. Esegui: pip install openpyxl",
            file=sys.stderr,
        )
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eval Classifier"
    ws.append(CSV_COLUMNS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    col_index = {col: idx + 1 for idx, col in enumerate(CSV_COLUMNS)}

    for row_data in results:
        ws.append([row_data.get(col, "") for col in CSV_COLUMNS])
        excel_row = ws.max_row

        for col in CORE_METRICS + ["query_clarity", "query_extractability"]:
            val = row_data.get(col)
            if isinstance(val, (int, float)) and 1 <= int(val) <= 5:
                ws.cell(row=excel_row, column=col_index[col]).fill = PatternFill(
                    fill_type="solid", fgColor=_SCORE_COLORS[int(val)]
                )

        fs = row_data.get("final_score")
        if isinstance(fs, (int, float)):
            ws.cell(row=excel_row, column=col_index["final_score"]).fill = PatternFill(
                fill_type="solid", fgColor=_SCORE_COLORS[max(1, min(5, round(fs)))]
            )

        cls = str(row_data.get("classification") or "").upper()
        if cls in _CLASSIFICATION_COLORS:
            ws.cell(row=excel_row, column=col_index["classification"]).fill = PatternFill(
                fill_type="solid", fgColor=_CLASSIFICATION_COLORS[cls]
            )

        rc = str(row_data.get("query_root_cause") or "").upper()
        if rc in _ROOT_CAUSE_COLORS:
            ws.cell(row=excel_row, column=col_index["query_root_cause"]).fill = PatternFill(
                fill_type="solid", fgColor=_ROOT_CAUSE_COLORS[rc]
            )

    _WIDE = {
        "query", "response", "candidates_names",
        "classifier_strengths", "classifier_weaknesses", "improvement_suggestion",
        "query_eval_notes",
    }
    for idx, col in enumerate(CSV_COLUMNS, 1):
        letter = get_column_letter(idx)
        if col in _WIDE:
            ws.column_dimensions[letter].width = 50
        elif col.startswith("reasoning_"):
            ws.column_dimensions[letter].width = 40
        else:
            ws.column_dimensions[letter].width = 18

    xl_path = path.with_suffix(".xlsx")
    wb.save(xl_path)
    print(f"[DONE] Excel saved to {xl_path}")


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------
results: list[dict[str, Any]] = []
total = len(rows)

print(f"[INFO] Agent       : {AGENT_ID}")
print(f"[INFO] Model       : {FOUNDRY_MODEL}")
print(f"[INFO] Endpoint    : {PROJECT_ENDPOINT}")
print(f"[INFO] Judge model : {JUDGE_MODEL}")
print(f"[INFO] Output      : {output_path}")
print(f"[INFO] Righe totali: {total}")
print(f"[INFO] NOTE: Nessuna ground truth — valutazione basata su coerenza e pertinenza.")
print()

try:
    for i, row in enumerate(rows, 1):
        query = _text(row.get("query"))
        title = _text(row.get("title"))

        print(f"[{i:>3}/{total}] Agent... ", end="", flush=True)
        t0 = time.time()
        agent_result = call_agent(query)
        response     = agent_result["response_text"]
        latency      = round(time.time() - t0, 3)

        if agent_result["agent_error"]:
            print(f"({latency:.1f}s) [ERROR] {agent_result['agent_error'][:80]}")
        else:
            preview = response[:90].replace("\n", " ")
            print(f"({latency:.1f}s) => {preview!r}")

        print(f"         Judge...  ", end="", flush=True)
        t1 = time.time()
        if agent_result["agent_error"]:
            scores: dict[str, Any] = {
                "classification": "AGENT_ERROR",
                **{m: 1 for m in CORE_METRICS},
                "final_score": 1.0,
                "classifier_strengths":   "",
                "classifier_weaknesses":  agent_result["agent_error"],
                "improvement_suggestion": "Fix agent error.",
                **{f"reasoning_{m}": "" for m in CORE_METRICS},
            }
        else:
            structured_classification = json.dumps(
                {
                    "skills": agent_result.get("classified_skills", ""),
                    "role": agent_result.get("classified_role", ""),
                    "location": agent_result.get("classified_location", ""),
                    "seniority": agent_result.get("classified_seniority", ""),
                    "language": agent_result.get("classified_language", ""),
                    "work_mode": agent_result.get("classified_work_mode", ""),
                },
                ensure_ascii=False,
            )
            scores = judge_score(
                query=query,
                title=title,
                response=response,
                structured_classification=structured_classification,
                search_evidence=agent_result.get("search_index_evidence", "") or "{}",
            )
        judge_latency = round(time.time() - t1, 3)
        print(
            f"({judge_latency:.1f}s) "
            f"class={scores.get('classification')} "
            f"coherence={scores.get('classification_coherence')} "
            f"final={scores.get('final_score')}"
        )

        print(f"         QEval...  ", end="", flush=True)
        t2 = time.time()
        qeval = assess_query(query=query, title=title)
        print(
            f"({time.time()-t2:.1f}s) "
            f"clarity={qeval.get('query_clarity')} "
            f"extractability={qeval.get('query_extractability')} "
            f"root_cause={qeval.get('query_root_cause')}"
        )

        results.append({
            "query":   query,
            "title":   title,
            "response":    response,
            "agent_error": agent_result["agent_error"],
            # campi estratti
            "classified_skills":    agent_result["classified_skills"],
            "classified_role":      agent_result["classified_role"],
            "classified_location":  agent_result["classified_location"],
            "classified_seniority": agent_result["classified_seniority"],
            "classified_language":  agent_result["classified_language"],
            "classified_work_mode": agent_result["classified_work_mode"],
            "candidates_count":     agent_result["candidates_count"],
            "candidates_names":     agent_result["candidates_names"],
            "search_index_evidence": agent_result["search_index_evidence"],
            # timing
            "latency_seconds":       latency,
            "judge_latency_seconds": judge_latency,
            "response_id":           agent_result["response_id"],
            **scores,
            **qeval,
        })

finally:
    if results:
        with output_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(results)
        print(f"\n[DONE] {len(results)} righe scritte su {output_path}")
        _write_excel(results, output_path)

        avg_final = [r["final_score"] for r in results if isinstance(r.get("final_score"), (int, float))]
        if avg_final:
            print(f"[SUMMARY] avg final_score = {sum(avg_final)/len(avg_final):.3f} su {len(avg_final)} righe")

        # Riepilogo per categoria
        by_title: dict[str, list[float]] = {}
        for r in results:
            t = r.get("title", "")
            fs = r.get("final_score")
            if isinstance(fs, (int, float)):
                by_title.setdefault(t, []).append(float(fs))
        if by_title:
            print("\n[SUMMARY per categoria]")
            for cat, scores_list in sorted(by_title.items()):
                avg = sum(scores_list) / len(scores_list)
                print(f"  {cat:<20} avg={avg:.3f}  n={len(scores_list)}")
    else:
        print("[WARNING] Nessun risultato da scrivere.")
